from core.empresa import Empresa
from core.proyecto import Proyecto
from core.persona import Persona

from pathlib import Path
from openpyxl import load_workbook


def escribir_excel(empresa : Empresa, ruta_excel: Path, proyecto : Proyecto, persona_atiende : Persona, 
                   persona_t1 : Persona, persona_t2 : Persona , persona_representante : Persona):


    
    excel = load_workbook(filename=ruta_excel)
    if not "De" in excel.sheetnames:
        print("No extiste la hoja 'DE'")
        return False
    hoja = excel["De"]

    hoja["B3"] = empresa.razon_social_propietario
    hoja["B4"] = empresa.razon_social_usuario
    hoja["B6"] = empresa.rfc
    hoja["B7"] = empresa.domicilio
    hoja["B11"] = empresa.telefono
    hoja["B12"] = empresa.empresa_a_facturar
    hoja["B13"] = proyecto.numero_equipos
    hoja["B14"] = empresa.actividad_economica
    hoja["B8"] = persona_atiende.nombre_completo
    hoja["B9"] = persona_t1.nombre_completo
    hoja["B10"] = persona_t2.nombre_completo
    hoja["B5"] =persona_representante.nombre_completo

    excel.save(ruta_excel)
    